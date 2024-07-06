from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
#1 - Lê pasta de trabalho e planilha
#importando workbook
wb = load_workbook('Automa_info\data\pivot_table.xlsx')
#acessando a planilha relatório
sheet= wb['Relatorio']

#2 - Referências das linhas e colunas
#capturando o min e max de colunas na planilha
min_column = wb.active.min_column
max_column = wb.active.max_column
#capturando o min e max de linhas na planilha
min_row = wb.active.min_row
max_row = wb.active.max_row

#3- adicionando dados e categorias no gráfico
barchart = BarChart()

data = Reference(
    sheet,
    min_col = min_column+1,
    max_col = max_column,
    min_row  = min_row,
    max_row = max_row
)

categories = Reference(
    sheet,
    min_col = min_column,
    max_col = min_column,
    min_row = min_row+1,
    max_row = max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#4 - criando o gráfico
sheet.add_chart(barchart, "B10")
barchart.title = 'Vendas por Fabricantes'
barchart.style = 2

#5 - salvando o workbook
wb.save("Automa_info\chart.xlsx")